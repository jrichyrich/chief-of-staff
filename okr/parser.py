"""Parse OKR data from an Excel spreadsheet.

Uses header-based column mapping instead of hardcoded indices so that
inserting or reordering columns in Excel does not silently corrupt data.
"""
from datetime import datetime
from pathlib import Path

import openpyxl

from okr.models import Initiative, KeyResult, OKRSnapshot, Objective


# ---------------------------------------------------------------------------
# Required tabs
# ---------------------------------------------------------------------------
REQUIRED_TABS = {"Objectives", "Key Results", "Initiatives"}

# ---------------------------------------------------------------------------
# Column definitions per tab: field_name → list of header aliases (lowercased)
# First alias is the canonical name in the current spreadsheet.
# ---------------------------------------------------------------------------
OBJECTIVE_COLUMNS = {
    "okr_id": ["okr id"],
    "name": ["objective name", "name"],
    "statement": ["objective statement", "statement"],
    "owner": ["owner (person)", "owner"],
    "team": ["owning team", "team"],
    "year": ["year"],
    "status": ["status"],
    "pct_complete": ["% complete"],
}

KEY_RESULT_COLUMNS = {
    "kr_id": ["kr id"],
    "okr_id": ["okr id"],
    "name": ["key result name", "name"],
    "status": ["status"],
    "pct_complete": ["% complete"],
    "baseline": ["baseline value", "baseline"],
    "target": ["2026 target value", "target value", "target"],
    "owner": ["owner (person)", "owner"],
    "team": ["owning team", "team"],
    "q1_milestone": ["q1 milestone"],
    "q2_milestone": ["q2 milestone"],
    "q3_milestone": ["q3 milestone"],
    "q4_milestone": ["q4 milestone"],
    "current_actual": ["current actual"],
    "gap_to_target": ["gap to target"],
}

INITIATIVE_COLUMNS = {
    "initiative_id": ["initiative id"],
    "kr_ids": ["kr id(s)", "kr ids"],
    "okr_id": ["okr id"],
    "name": ["initiative name", "name"],
    "pct_complete": ["% complete"],
    "status": ["status"],
    "blocker": ["blocker"],
    "description": ["description"],
    "investment_tier": ["investment tier"],
    "maturity_type": ["maturity type"],
    "owner": ["owner (person)", "owner"],
    "team": ["owning team", "team"],
    "timeline_start": ["timeline start"],
    "timeline_end": ["timeline end"],
    "investment_dollars": ["investment ($)", "investment dollars"],
}

# Required fields per tab (must be present or parsing fails)
REQUIRED_OBJECTIVE_FIELDS = {
    "okr_id", "name", "statement", "owner", "team", "year", "status",
    "pct_complete",
}
REQUIRED_KEY_RESULT_FIELDS = {"kr_id", "okr_id", "name", "status", "pct_complete"}
REQUIRED_INITIATIVE_FIELDS = {"initiative_id", "kr_ids", "okr_id", "name"}


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------
def _cell_str(value) -> str:
    """Convert a cell value to a clean string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _cell_float(value) -> float:
    """Convert a cell value to a float, defaulting to 0.0."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _cell_pct(value) -> float:
    """Convert an Excel decimal percentage to a display percentage.

    Excel stores percentages as decimals (0.5 = 50%).  This function
    multiplies by 100 so the stored value reads as a human-friendly
    percentage (50.0 instead of 0.5).
    """
    return round(_cell_float(value) * 100, 2)


# ---------------------------------------------------------------------------
# Column mapping helpers
# ---------------------------------------------------------------------------
def _build_column_map(header_row) -> dict[str, int]:
    """Build a lowercased, stripped column-name-to-index mapping from a header row."""
    return {
        str(cell.value).strip().lower(): idx
        for idx, cell in enumerate(header_row)
        if cell.value is not None
    }


def _resolve_columns(
    col_map: dict[str, int],
    column_defs: dict[str, list[str]],
    required_fields: set[str],
    tab_name: str,
) -> dict[str, int]:
    """Resolve field names to column indices using the header map.

    For each field in *column_defs*, tries each alias in order and maps the
    field to the first matching header index.

    Raises ``ValueError`` if any *required_fields* cannot be resolved.
    """
    resolved: dict[str, int] = {}
    for field_name, aliases in column_defs.items():
        for alias in aliases:
            if alias in col_map:
                resolved[field_name] = col_map[alias]
                break

    missing = required_fields - set(resolved.keys())
    if missing:
        raise ValueError(
            f"Missing required columns in {tab_name} tab: {missing}. "
            f"Available headers: {sorted(col_map.keys())}"
        )
    return resolved


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def parse_okr_spreadsheet(path: Path) -> OKRSnapshot:
    """Parse an OKR Excel spreadsheet and return an OKRSnapshot.

    Args:
        path: Path to the .xlsx file.

    Returns:
        OKRSnapshot with parsed objectives, key results, and initiatives.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If required tabs or columns are missing.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"OKR spreadsheet not found: {path}")

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

    try:
        # Validate required tabs
        available_tabs = set(wb.sheetnames)
        missing_tabs = REQUIRED_TABS - available_tabs
        if missing_tabs:
            raise ValueError(
                f"Missing required tabs in OKR spreadsheet: {missing_tabs}"
            )

        objectives = _parse_objectives(wb["Objectives"])
        key_results = _parse_key_results(wb["Key Results"])
        initiatives = _parse_initiatives(wb["Initiatives"])
    finally:
        wb.close()

    return OKRSnapshot(
        timestamp=datetime.now().isoformat(),
        source_file=str(path),
        objectives=objectives,
        key_results=key_results,
        initiatives=initiatives,
    )


# ---------------------------------------------------------------------------
# Tab parsers
# ---------------------------------------------------------------------------
def _parse_objectives(ws) -> list[Objective]:
    """Parse the Objectives tab using header-based column mapping."""
    rows = ws.iter_rows()
    header_row = next(rows)
    col_map = _build_column_map(header_row)
    cols = _resolve_columns(
        col_map, OBJECTIVE_COLUMNS, REQUIRED_OBJECTIVE_FIELDS, "Objectives",
    )

    objectives = []
    for row in rows:
        cells = [c.value for c in row]
        if cells[cols["okr_id"]] is None:
            continue
        objectives.append(
            Objective(
                okr_id=_cell_str(cells[cols["okr_id"]]),
                name=_cell_str(cells[cols["name"]]),
                statement=_cell_str(cells[cols["statement"]]),
                owner=_cell_str(cells[cols["owner"]]),
                team=_cell_str(cells[cols["team"]]),
                year=_cell_str(cells[cols["year"]]),
                status=_cell_str(cells[cols["status"]]),
                pct_complete=_cell_pct(cells[cols["pct_complete"]]),
            )
        )
    return objectives


def _parse_key_results(ws) -> list[KeyResult]:
    """Parse the Key Results tab using header-based column mapping."""
    rows = ws.iter_rows()
    header_row = next(rows)
    col_map = _build_column_map(header_row)
    cols = _resolve_columns(
        col_map, KEY_RESULT_COLUMNS, REQUIRED_KEY_RESULT_FIELDS, "Key Results",
    )

    key_results = []
    for row in rows:
        cells = [c.value for c in row]
        if cells[cols["kr_id"]] is None:
            continue
        key_results.append(
            KeyResult(
                kr_id=_cell_str(cells[cols["kr_id"]]),
                okr_id=_cell_str(cells[cols["okr_id"]]),
                name=_cell_str(cells[cols["name"]]),
                status=_cell_str(cells[cols["status"]]),
                pct_complete=_cell_pct(cells[cols["pct_complete"]]),
                baseline=_cell_str(cells[cols["baseline"]]) if "baseline" in cols else "",
                target=_cell_str(cells[cols["target"]]) if "target" in cols else "",
                owner=_cell_str(cells[cols["owner"]]) if "owner" in cols else "",
                team=_cell_str(cells[cols["team"]]) if "team" in cols else "",
                q1_milestone=_cell_str(cells[cols["q1_milestone"]]) if "q1_milestone" in cols else "",
                q2_milestone=_cell_str(cells[cols["q2_milestone"]]) if "q2_milestone" in cols else "",
                q3_milestone=_cell_str(cells[cols["q3_milestone"]]) if "q3_milestone" in cols else "",
                q4_milestone=_cell_str(cells[cols["q4_milestone"]]) if "q4_milestone" in cols else "",
                current_actual=_cell_str(cells[cols["current_actual"]]) if "current_actual" in cols else "",
                gap_to_target=_cell_str(cells[cols["gap_to_target"]]) if "gap_to_target" in cols else "",
            )
        )
    return key_results


def _parse_initiatives(ws) -> list[Initiative]:
    """Parse the Initiatives tab using header-based column mapping."""
    rows = ws.iter_rows()
    header_row = next(rows)
    col_map = _build_column_map(header_row)
    cols = _resolve_columns(
        col_map, INITIATIVE_COLUMNS, REQUIRED_INITIATIVE_FIELDS, "Initiatives",
    )

    initiatives = []
    for row in rows:
        cells = [c.value for c in row]
        if cells[cols["initiative_id"]] is None:
            continue
        initiatives.append(
            Initiative(
                initiative_id=_cell_str(cells[cols["initiative_id"]]),
                kr_ids=_cell_str(cells[cols["kr_ids"]]),
                okr_id=_cell_str(cells[cols["okr_id"]]),
                name=_cell_str(cells[cols["name"]]),
                pct_complete=_cell_pct(cells[cols["pct_complete"]]) if "pct_complete" in cols else 0.0,
                status=_cell_str(cells[cols["status"]]) if "status" in cols else "",
                blocker=_cell_str(cells[cols["blocker"]]) if "blocker" in cols else "",
                description=_cell_str(cells[cols["description"]]) if "description" in cols else "",
                investment_tier=_cell_str(cells[cols["investment_tier"]]) if "investment_tier" in cols else "",
                maturity_type=_cell_str(cells[cols["maturity_type"]]) if "maturity_type" in cols else "",
                owner=_cell_str(cells[cols["owner"]]) if "owner" in cols else "",
                team=_cell_str(cells[cols["team"]]) if "team" in cols else "",
                timeline_start=_cell_str(cells[cols["timeline_start"]]) if "timeline_start" in cols else "",
                timeline_end=_cell_str(cells[cols["timeline_end"]]) if "timeline_end" in cols else "",
                investment_dollars=_cell_float(cells[cols["investment_dollars"]]) if "investment_dollars" in cols else 0.0,
            )
        )
    return initiatives
